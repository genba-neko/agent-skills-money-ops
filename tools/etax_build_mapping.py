"""NTA e-Tax XML 構造設計書 (Excel) からフィールドコード→項目名 mapping JSON を生成。

入力 (一時、git 対象外):
    .work/etax_spec/extracted/09XML構造設計書等【所得税】/XML構造設計書(所得-申告)Ver*.xlsx

出力:
    src/money_ops/etax/mapping/<form>.json
    例: KOA020.json = {"_meta": {...}, "fields": {"ABA00010": "年分", ...}}

各帳票 (KOA020 等) は複数の Ver で更新されている。
最新 Ver を採用 (申告書面の改訂を反映)。

NTA 公式仕様書の利用条件: 商用・非商用利用可、財務会計ソフト開発に使用可。
本スクリプト出力は (フィールドコード, 短いラベル) の対応のみ → 事実情報。

Excel 列 (XML 構造設計書):
  c0  項番
  c1  レベル (階層深さ)
  c4-c11  階層別項目名 (人間可読)
  c12  入力型 (kubun/zipcode/address/name 等)
  c16  ENUM 名
  c17  page (帳票識別)
  c18  フィールドコード ★

使い方:
    python tools/etax_build_mapping.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
_SPEC_DIR = _PROJECT_ROOT / ".work" / "etax_spec" / "extracted" / "09XML構造設計書等【所得税】"
_OUT_DIR = _PROJECT_ROOT / "src" / "money_ops" / "etax" / "mapping"

_HIER_COLS = list(range(4, 12))   # c4-c11 階層別項目名
_FIELD_TYPE_COL = 12              # 入力型
_ENUM_COL = 16                    # ENUM 名
_FIELD_CODE_COL = 18              # フィールドコード


def _rightmost_label(row: tuple) -> str:
    """c4-c11 の最右非空セルを取得 = 最深階層の項目名。"""
    for ci in reversed(_HIER_COLS):
        if ci < len(row) and row[ci] is not None:
            s = str(row[ci]).strip()
            if s:
                return s
    return ""


def _hier_path(row: tuple) -> list[str]:
    """c4-c11 の階層パス (root → leaf 順、空セル除外)。"""
    path = []
    for ci in _HIER_COLS:
        if ci < len(row) and row[ci] is not None:
            s = str(row[ci]).strip()
            if s:
                path.append(s)
    return path


def _scan_form_sheet(ws) -> dict:
    """1帳票シートからフィールドコード→ラベル + メタ情報抽出。"""
    fields: dict = {}
    pages: set = set()
    for row in ws.iter_rows(values_only=True):
        if len(row) <= _FIELD_CODE_COL:
            continue
        code = row[_FIELD_CODE_COL]
        if not code:
            continue
        code_s = str(code).strip()
        # フィールドコード形式 (例 ABA00010, BFC00000) を持つ行のみ
        if not re.match(r"^[A-Z]{2,4}\d{5}$", code_s):
            # page (KOA020-1 等) は別管理
            if re.match(r"^[A-Z]{3,4}\d{3}-\d+$", code_s):
                pages.add(code_s)
            continue
        label = _rightmost_label(row)
        if not label:
            continue
        path = _hier_path(row)
        ftype = row[_FIELD_TYPE_COL] if len(row) > _FIELD_TYPE_COL and row[_FIELD_TYPE_COL] else ""
        enum = row[_ENUM_COL] if len(row) > _ENUM_COL and row[_ENUM_COL] else ""
        # 同じコードが複数行に登場する場合 (繰り返しグループ等) は最初を採用
        if code_s not in fields:
            fields[code_s] = {
                "label": label,
                "path": path,
                "type": str(ftype).strip() if ftype else "",
                "enum": str(enum).strip() if enum else "",
            }
    return {"fields": fields, "pages": sorted(pages)}


def main() -> None:
    if not _SPEC_DIR.exists():
        print(f"[ERR] 仕様書ディレクトリなし: {_SPEC_DIR}", file=sys.stderr)
        print("先に .work/etax_spec/ に e-tax09.CAB を展開してください", file=sys.stderr)
        sys.exit(1)

    spec_files = sorted(
        _SPEC_DIR.glob("XML構造設計書(所得-申告)Ver*.xlsx"),
        key=lambda p: int(re.search(r"Ver(\d+)x", p.name).group(1)),
    )
    if not spec_files:
        print(f"[ERR] Ver xlsx が見つからない", file=sys.stderr)
        sys.exit(1)

    # 各帳票が最後に登場した Ver の sheet を採用
    form_to_latest: dict[str, tuple[int, Path]] = {}
    for p in spec_files:
        ver = int(re.search(r"Ver(\d+)x", p.name).group(1))
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        for sname in wb.sheetnames:
            form_to_latest[sname] = (ver, p)
        wb.close()

    print(f"対象 Ver xlsx: {len(spec_files)} 個")
    print(f"対象帳票     : {len(form_to_latest)} 個")
    print(f"出力先       : {_OUT_DIR}")
    _OUT_DIR.mkdir(parents=True, exist_ok=True)

    total_fields = 0
    for form, (ver, xl_path) in sorted(form_to_latest.items()):
        wb = openpyxl.load_workbook(xl_path, read_only=True, data_only=True)
        ws = wb[form]
        result = _scan_form_sheet(ws)
        wb.close()
        n = len(result["fields"])
        if n == 0:
            print(f"  [SKIP] {form} (Ver{ver}, fields=0)")
            continue
        out = {
            "_meta": {
                "form_code": form,
                "source_ver": f"Ver{ver}",
                "source": "NTA e-tax09.CAB / XML構造設計書(所得-申告)",
                "tax_type": "shotoku",
                "pages": result["pages"],
            },
            "fields": result["fields"],
        }
        out_path = _OUT_DIR / f"{form}.json"
        out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        total_fields += n
        print(f"  [OK]   {form:<10} Ver{ver:>2}  fields={n:>4}  pages={result['pages']}")

    print(f"\n合計 mapping fields: {total_fields}")


if __name__ == "__main__":
    main()
